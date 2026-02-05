#!/usr/bin/env python3
"""
docdiff.py - v0.1 Construction PDF set differ (text+tables+basic spec sectioning) -> Excel Change Queue

Usage:
  python docdiff.py --gmp ./input/GMP --bid ./input/BID --addenda ./input/ADDENDA --out ./output/changes.xlsx --config ./config.yaml
"""

from __future__ import annotations

import argparse
import dataclasses
import hashlib
import os
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import fitz  # PyMuPDF
import pdfplumber
import yaml
from rapidfuzz import fuzz, process
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# -------------------------
# Data models
# -------------------------

@dataclass
class PageExtract:
    pdf_path: str
    page_num: int
    text: str
    sheet_id: Optional[str]
    sheet_title_hint: Optional[str]
    tables: List[List[List[str]]]  # list of tables -> rows -> cells


@dataclass
class DocSet:
    name: str
    root: str
    pages: List[PageExtract]


@dataclass
class ChangeRow:
    change_id: str
    set_from: str
    set_to: str
    discipline: str
    doc_type: str
    reference: str
    change_type: str
    change_summary: str
    before_snippet: str
    after_snippet: str
    confidence: str
    auto_flags: str
    impact_score: int


# -------------------------
# Helpers
# -------------------------

def load_config(path: str) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def list_pdfs(folder: str) -> List[str]:
    p = Path(folder)
    if not p.exists():
        return []
    pdfs = sorted([str(x) for x in p.rglob("*.pdf")])
    return pdfs


def normalize_whitespace(s: str) -> str:
    s = s.replace("\r", "\n")
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()


def short_hash(*parts: str) -> str:
    h = hashlib.sha1()
    for p in parts:
        h.update(p.encode("utf-8", errors="ignore"))
        h.update(b"\0")
    return h.hexdigest()[:10]


def guess_discipline(sheet_id: Optional[str]) -> str:
    if not sheet_id:
        return "Unknown"
    s = sheet_id.upper()
    # Common discipline letters
    if s.startswith("A"):
        return "Architectural"
    if s.startswith("S"):
        return "Structural"
    if s.startswith("M") or s.startswith("ME") or s.startswith("MP"):
        return "Mechanical"
    if s.startswith("P") or s.startswith("PL"):
        return "Plumbing"
    if s.startswith("E") or s.startswith("EL"):
        return "Electrical"
    if s.startswith("FP") or s.startswith("FA"):
        return "Fire Protection"
    if s.startswith("C"):
        return "Civil"
    if s.startswith("L"):
        return "Landscape"
    return "Unknown"


def apply_flags(config: Dict[str, Any], text: str) -> List[str]:
    flags = []
    t = text.lower()
    for group, terms in (config.get("flags") or {}).items():
        for term in terms:
            if term.lower() in t:
                flags.append(group)
                break
    return flags


def compute_impact_score(change_type: str, flags: List[str], before: str, after: str, is_new_section: bool = False, table_delta: bool = False) -> int:
    score = 0
    if is_new_section:
        score += 25
    if table_delta:
        score += 20
    if change_type.lower() == "added":
        score += 10
    if change_type.lower() == "removed":
        score += 8
    if flags:
        score += 15
    # "provide" responsibility language
    if re.search(r"\bprovide\b", after.lower()) and not re.search(r"\bprovide\b", before.lower()):
        score += 10
    return min(score, 100)


# -------------------------
# Extraction
# -------------------------

def extract_tables_pdfplumber(pdf_path: str, page_num: int) -> List[List[List[str]]]:
    """
    Extract tables as list-of-tables, each table = rows, each row = cells (strings).
    This works reasonably on selectable-text schedule tables; it won't catch every drawn-line grid.
    """
    tables: List[List[List[str]]] = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            page = pdf.pages[page_num]
            raw_tables = page.extract_tables()
            for t in raw_tables or []:
                cleaned = []
                for row in t:
                    if row is None:
                        continue
                    cleaned.append([(c or "").strip() for c in row])
                # drop tiny/noisy tables
                if len(cleaned) >= 3 and max(len(r) for r in cleaned) >= 2:
                    tables.append(cleaned)
    except Exception:
        return []
    return tables


def find_sheet_id(config: Dict[str, Any], text: str) -> Optional[str]:
    patterns = config.get("sheet_id_patterns") or []
    candidates = []
    for pat in patterns:
        for m in re.finditer(pat, text, flags=re.IGNORECASE):
            candidates.append(m.group(0))
    if not candidates:
        return None
    # Prefer formats with dash (A-101) or letter+digits
    def score(c: str) -> int:
        c2 = c.upper().replace(" ", "")
        sc = 0
        if "-" in c2:
            sc += 5
        if re.match(r"^[A-Z]{1,3}-?\d", c2):
            sc += 3
        sc += min(len(c2), 10)
        return sc
    best = sorted(candidates, key=score, reverse=True)[0]
    return best.upper().replace(" ", "")


def title_hint_from_text(text: str, sheet_id: Optional[str]) -> Optional[str]:
    if not text:
        return None
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    if not lines:
        return None
    # heuristic: look for a line near the sheet id, otherwise first strong line
    if sheet_id:
        for i, line in enumerate(lines[:40]):
            if sheet_id in line.replace(" ", "").upper():
                # next 1-3 lines might be title
                for j in range(i + 1, min(i + 4, len(lines))):
                    cand = lines[j]
                    if 6 <= len(cand) <= 80 and not re.match(r"^\d+$", cand):
                        return cand[:80]
    # fallback: first line that looks like a title
    for line in lines[:30]:
        if 8 <= len(line) <= 80 and not re.search(r"\b(issued|revision|date|scale)\b", line, re.IGNORECASE):
            return line[:80]
    return None


def extract_pdf_pages(config: Dict[str, Any], pdf_path: str) -> List[PageExtract]:
    pages: List[PageExtract] = []
    doc = fitz.open(pdf_path)
    for i in range(doc.page_count):
        page = doc.load_page(i)
        text = page.get_text("text") or ""
        text = normalize_whitespace(text)
        sheet_id = find_sheet_id(config, text)
        title_hint = title_hint_from_text(text, sheet_id)
        tables = extract_tables_pdfplumber(pdf_path, i)
        pages.append(PageExtract(
            pdf_path=pdf_path,
            page_num=i,
            text=text,
            sheet_id=sheet_id,
            sheet_title_hint=title_hint,
            tables=tables
        ))
    doc.close()
    return pages


def ingest_set(config: Dict[str, Any], name: str, folder: str) -> DocSet:
    pdfs = list_pdfs(folder)
    pages: List[PageExtract] = []
    for pdf in pdfs:
        pages.extend(extract_pdf_pages(config, pdf))
    return DocSet(name=name, root=folder, pages=pages)


# -------------------------
# Matching
# -------------------------

def build_sheet_index(docset: DocSet) -> Dict[str, List[PageExtract]]:
    idx: Dict[str, List[PageExtract]] = {}
    for p in docset.pages:
        key = p.sheet_id or f"UNKEYED::{Path(p.pdf_path).name}::p{p.page_num+1}"
        idx.setdefault(key, []).append(p)
    return idx


def match_sheet_keys(gmp_keys: List[str], bid_keys: List[str]) -> Dict[str, Optional[str]]:
    """
    Map GMP sheet keys -> BID sheet keys.
    Primary: exact match on sheet id.
    Fallback: fuzzy match on key strings.
    """
    bid_set = set(bid_keys)
    mapping: Dict[str, Optional[str]] = {}
    for k in gmp_keys:
        if k in bid_set:
            mapping[k] = k
        else:
            # fuzzy match on keys that look like sheet ids
            choices = [b for b in bid_keys if not b.startswith("UNKEYED::")]
            if choices:
                best = process.extractOne(k, choices, scorer=fuzz.WRatio)
                if best and best[1] >= 90:
                    mapping[k] = best[0]
                else:
                    mapping[k] = None
            else:
                mapping[k] = None
    return mapping


# -------------------------
# Diff logic (notes + tables + specs-basic)
# -------------------------

def split_note_bullets(config: Dict[str, Any], text: str) -> List[str]:
    """
    Extract bullet-ish/numbered note lines.
    This is intentionally conservative: better to miss some than flood noise.
    """
    min_len = int((config.get("diff") or {}).get("note_bullet_min_len", 12))
    lines = [l.strip() for l in text.splitlines()]
    bullets = []
    for l in lines:
        if len(l) < min_len:
            continue
        if re.match(r"^(\(?[0-9]{1,3}\)?[.)]|[A-Z][.)]|[-•])\s+", l):
            bullets.append(l)
        elif re.match(r"^(KEYNOTE|NOTE)\b", l, flags=re.IGNORECASE):
            bullets.append(l)
    # de-dupe while preserving order
    seen = set()
    out = []
    for b in bullets:
        k = re.sub(r"\s+", " ", b).strip().lower()
        if k not in seen:
            seen.add(k)
            out.append(b)
    return out


def diff_lists(before: List[str], after: List[str]) -> Tuple[List[str], List[str]]:
    bset = set([x.strip() for x in before if x.strip()])
    aset = set([x.strip() for x in after if x.strip()])
    added = sorted(list(aset - bset))
    removed = sorted(list(bset - aset))
    return added, removed


def table_signature(table: List[List[str]]) -> List[str]:
    sig = []
    for row in table:
        row_join = " | ".join([c.strip() for c in row if c is not None]).strip()
        row_join = re.sub(r"\s+", " ", row_join)
        if row_join:
            sig.append(row_join)
    return sig


def diff_tables(before_tables: List[List[List[str]]], after_tables: List[List[List[str]]]) -> Tuple[int, int, int]:
    """
    Return rough counts: rows_added, rows_removed, rows_changed (by signature).
    """
    b_rows = []
    a_rows = []
    for t in before_tables:
        b_rows.extend(table_signature(t))
    for t in after_tables:
        a_rows.extend(table_signature(t))
    bset = set(b_rows)
    aset = set(a_rows)
    rows_added = len(aset - bset)
    rows_removed = len(bset - aset)
    # changed is hard without identity; approximate as overlap mismatch
    rows_changed = 0
    return rows_added, rows_removed, rows_changed


def extract_spec_sections(config: Dict[str, Any], text: str) -> Dict[str, str]:
    """
    Basic spec section splitter based on "SECTION XX XX XX".
    If not found, returns single bucket "UNKNOWN".
    """
    patterns = config.get("spec_section_patterns") or []
    if not patterns:
        return {"UNKNOWN": text}

    # Find all section starts
    starts: List[Tuple[int, str]] = []
    for pat in patterns:
        for m in re.finditer(pat, text, flags=re.IGNORECASE):
            sec = m.group(1) if m.lastindex else m.group(0)
            sec = re.sub(r"\s+", " ", sec).strip()
            starts.append((m.start(), sec))
    if not starts:
        return {"UNKNOWN": text}

    starts = sorted(starts, key=lambda x: x[0])
    sections: Dict[str, str] = {}
    for idx, (pos, sec) in enumerate(starts):
        end = starts[idx + 1][0] if idx + 1 < len(starts) else len(text)
        chunk = text[pos:end].strip()
        if sec not in sections or len(chunk) > len(sections[sec]):
            sections[sec] = chunk
    return sections


# -------------------------
# Change generation
# -------------------------

def make_change_id(set_from: str, set_to: str, reference: str, change_type: str, summary: str) -> str:
    return short_hash(set_from, set_to, reference, change_type, summary)


def build_inventory_changes(gmp: DocSet, bid: DocSet) -> List[ChangeRow]:
    g_idx = build_sheet_index(gmp)
    b_idx = build_sheet_index(bid)

    g_keys = set(g_idx.keys())
    b_keys = set(b_idx.keys())

    added_sheets = sorted(list(b_keys - g_keys))
    removed_sheets = sorted(list(g_keys - b_keys))

    rows: List[ChangeRow] = []
    for k in added_sheets:
        discipline = guess_discipline(k if not k.startswith("UNKEYED::") else None)
        summary = f"Sheet present in {bid.name} but not in {gmp.name}: {k}"
        rows.append(ChangeRow(
            change_id=make_change_id(gmp.name, bid.name, k, "Added", summary),
            set_from=gmp.name,
            set_to=bid.name,
            discipline=discipline,
            doc_type="Drawing",
            reference=k,
            change_type="Added",
            change_summary=summary,
            before_snippet="",
            after_snippet="",
            confidence="High",
            auto_flags="",
            impact_score=25
        ))
    for k in removed_sheets:
        discipline = guess_discipline(k if not k.startswith("UNKEYED::") else None)
        summary = f"Sheet present in {gmp.name} but not in {bid.name}: {k}"
        rows.append(ChangeRow(
            change_id=make_change_id(gmp.name, bid.name, k, "Removed", summary),
            set_from=gmp.name,
            set_to=bid.name,
            discipline=discipline,
            doc_type="Drawing",
            reference=k,
            change_type="Removed",
            change_summary=summary,
            before_snippet="",
            after_snippet="",
            confidence="High",
            auto_flags="",
            impact_score=20
        ))
    return rows


def compare_sets(config: Dict[str, Any], set_from: DocSet, set_to: DocSet) -> List[ChangeRow]:
    """
    Compare two sets and generate change rows for:
      - Note bullets
      - Tables (schedule-ish deltas)
      - Spec sections (if specs PDFs are included in the set)
    """
    rows: List[ChangeRow] = []

    f_idx = build_sheet_index(set_from)
    t_idx = build_sheet_index(set_to)

    f_keys = list(f_idx.keys())
    t_keys = list(t_idx.keys())

    mapping = match_sheet_keys(f_keys, t_keys)

    max_per_sheet = int((config.get("diff") or {}).get("max_changes_per_sheet", 200))
    max_snip = int((config.get("diff") or {}).get("max_snippet_chars", 700))

    for f_key, t_key in mapping.items():
        if not t_key:
            continue
        f_pages = f_idx.get(f_key, [])
        t_pages = t_idx.get(t_key, [])
        if not f_pages or not t_pages:
            continue

        # Many PDFs have 1 page per sheet; if multiple, concatenate text + tables
        f_text = "\n\n".join([p.text for p in f_pages if p.text])
        t_text = "\n\n".join([p.text for p in t_pages if p.text])

        f_tables = []
        t_tables = []
        for p in f_pages:
            f_tables.extend(p.tables or [])
        for p in t_pages:
            t_tables.extend(p.tables or [])

        # Determine doc_type: if looks like specs, treat as Spec
        is_spec_like = bool(re.search(r"\bSECTION\s+\d{2}\s+\d{2}\s+\d{2}\b", t_text, flags=re.IGNORECASE))

        reference = f_key if not f_key.startswith("UNKEYED::") else t_key
        discipline = guess_discipline(reference if not reference.startswith("UNKEYED::") else None)

        if is_spec_like:
            # Spec diff by sections
            f_secs = extract_spec_sections(config, f_text)
            t_secs = extract_spec_sections(config, t_text)

            f_sec_keys = set(f_secs.keys())
            t_sec_keys = set(t_secs.keys())

            for sec in sorted(list(t_sec_keys - f_sec_keys)):
                after = t_secs[sec]
                flags = apply_flags(config, after)
                summary = f"Spec section appears in {set_to.name} but not in {set_from.name}: {sec}"
                rows.append(ChangeRow(
                    change_id=make_change_id(set_from.name, set_to.name, f"Spec {sec}", "Added", summary),
                    set_from=set_from.name,
                    set_to=set_to.name,
                    discipline="Specifications",
                    doc_type="Spec",
                    reference=f"{sec}",
                    change_type="Added",
                    change_summary=summary,
                    before_snippet="",
                    after_snippet=after[:max_snip],
                    confidence="High",
                    auto_flags=";".join(flags),
                    impact_score=compute_impact_score("Added", flags, "", after, is_new_section=True)
                ))

            for sec in sorted(list(f_sec_keys - t_sec_keys)):
                before = f_secs[sec]
                summary = f"Spec section present in {set_from.name} but missing in {set_to.name}: {sec}"
                rows.append(ChangeRow(
                    change_id=make_change_id(set_from.name, set_to.name, f"Spec {sec}", "Removed", summary),
                    set_from=set_from.name,
                    set_to=set_to.name,
                    discipline="Specifications",
                    doc_type="Spec",
                    reference=f"{sec}",
                    change_type="Removed",
                    change_summary=summary,
                    before_snippet=before[:max_snip],
                    after_snippet="",
                    confidence="High",
                    auto_flags="",
                    impact_score=20
                ))

            # Modified sections: compare hash of normalized chunks
            for sec in sorted(list(f_sec_keys & t_sec_keys)):
                before = f_secs[sec]
                after = t_secs[sec]
                if short_hash(normalize_whitespace(before)) != short_hash(normalize_whitespace(after)):
                    flags = apply_flags(config, after)
                    summary = f"Spec section modified: {sec}"
                    rows.append(ChangeRow(
                        change_id=make_change_id(set_from.name, set_to.name, f"Spec {sec}", "Modified", summary),
                        set_from=set_from.name,
                        set_to=set_to.name,
                        discipline="Specifications",
                        doc_type="Spec",
                        reference=f"{sec}",
                        change_type="Modified",
                        change_summary=summary,
                        before_snippet=before[:max_snip],
                        after_snippet=after[:max_snip],
                        confidence="High",
                        auto_flags=";".join(flags),
                        impact_score=compute_impact_score("Modified", flags, before, after)
                    ))
            continue

        # Drawing-like content: notes & tables
        f_bul = split_note_bullets(config, f_text)
        t_bul = split_note_bullets(config, t_text)
        added, removed = diff_lists(f_bul, t_bul)

        # emit a few aggregated rows rather than thousands of micro-rows
        if added:
            after = "\n".join(added[:max_per_sheet])
            flags = apply_flags(config, after)
            summary = f"Added note bullets on {reference}: {len(added)} item(s)"
            rows.append(ChangeRow(
                change_id=make_change_id(set_from.name, set_to.name, reference, "Added", summary),
                set_from=set_from.name,
                set_to=set_to.name,
                discipline=discipline,
                doc_type="Drawing",
                reference=reference,
                change_type="Added",
                change_summary=summary,
                before_snippet="",
                after_snippet=after[:max_snip],
                confidence="High" if len(added) <= 50 else "Med",
                auto_flags=";".join(flags),
                impact_score=compute_impact_score("Added", flags, "", after)
            ))

        if removed:
            before = "\n".join(removed[:max_per_sheet])
            summary = f"Removed note bullets on {reference}: {len(removed)} item(s)"
            rows.append(ChangeRow(
                change_id=make_change_id(set_from.name, set_to.name, reference, "Removed", summary),
                set_from=set_from.name,
                set_to=set_to.name,
                discipline=discipline,
                doc_type="Drawing",
                reference=reference,
                change_type="Removed",
                change_summary=summary,
                before_snippet=before[:max_snip],
                after_snippet="",
                confidence="High" if len(removed) <= 50 else "Med",
                auto_flags="",
                impact_score=compute_impact_score("Removed", [], before, "")
            ))

        # Tables: just detect row adds/removes (good for schedules)
        if f_tables or t_tables:
            rows_added, rows_removed, _ = diff_tables(f_tables, t_tables)
            if rows_added or rows_removed:
                after_sig = ""
                before_sig = ""
                # include a small signature sample
                if t_tables:
                    after_sig = "\n".join(table_signature(t_tables[0])[:40])[:max_snip]
                if f_tables:
                    before_sig = "\n".join(table_signature(f_tables[0])[:40])[:max_snip]

                table_text_for_flags = (after_sig or "") + "\n" + (before_sig or "")
                flags = apply_flags(config, table_text_for_flags)
                summary = f"Table/schedule delta on {reference}: +{rows_added} row(s), -{rows_removed} row(s)"
                rows.append(ChangeRow(
                    change_id=make_change_id(set_from.name, set_to.name, reference, "Modified", summary),
                    set_from=set_from.name,
                    set_to=set_to.name,
                    discipline=discipline,
                    doc_type="Drawing",
                    reference=reference,
                    change_type="Modified",
                    change_summary=summary,
                    before_snippet=before_sig,
                    after_snippet=after_sig,
                    confidence="Med",
                    auto_flags=";".join(flags),
                    impact_score=compute_impact_score("Modified", flags, before_sig, after_sig, table_delta=True)
                ))

    return rows


# -------------------------
# Excel output
# -------------------------

CHANGE_QUEUE_HEADERS = [
    "Change_ID",
    "Set_From",
    "Set_To",
    "Discipline",
    "Doc_Type",
    "Reference",
    "Change_Type",
    "Change_Summary",
    "Before_Snippet",
    "After_Snippet",
    "Confidence",
    "Auto_Flags",
    "Impact_Score",
    "Estimator_Significance_1to5",
    "Disposition",
    "Notes",
]

def autosize_columns(ws: Worksheet, max_width: int = 60) -> None:
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            val = str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_width)


def write_excel(config: Dict[str, Any], out_path: str, change_rows: List[ChangeRow], inv_rows: List[ChangeRow]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Change_Queue"
    ws.append(CHANGE_QUEUE_HEADERS)

    # Sort by Impact_Score desc then confidence
    conf_rank = {"High": 0, "Med": 1, "Low": 2}
    change_rows_sorted = sorted(
        change_rows,
        key=lambda r: (-r.impact_score, conf_rank.get(r.confidence, 9), r.discipline, r.reference)
    )

    for r in change_rows_sorted:
        ws.append([
            r.change_id,
            r.set_from,
            r.set_to,
            r.discipline,
            r.doc_type,
            r.reference,
            r.change_type,
            r.change_summary,
            r.before_snippet,
            r.after_snippet,
            r.confidence,
            r.auto_flags,
            r.impact_score,
            "",  # estimator significance
            "",  # disposition
            "",  # notes
        ])

    # Inventory tab
    ws2 = wb.create_sheet("Sheets_Inventory")
    ws2.append(CHANGE_QUEUE_HEADERS)
    for r in inv_rows:
        ws2.append([
            r.change_id, r.set_from, r.set_to, r.discipline, r.doc_type, r.reference,
            r.change_type, r.change_summary, r.before_snippet, r.after_snippet,
            r.confidence, r.auto_flags, r.impact_score, "", "", ""
        ])

    # Formatting
    if (config.get("excel") or {}).get("freeze_panes", True):
        ws.freeze_panes = "A2"
        ws2.freeze_panes = "A2"
    if (config.get("excel") or {}).get("auto_filter", True):
        ws.auto_filter.ref = f"A1:{get_column_letter(len(CHANGE_QUEUE_HEADERS))}1"
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(CHANGE_QUEUE_HEADERS))}1"

    autosize_columns(ws)
    autosize_columns(ws2)

    wb.save(out_path)


# -------------------------
# Main
# -------------------------

def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--gmp", required=True, help="Folder containing GMP PDFs")
    ap.add_argument("--bid", required=True, help="Folder containing Bid/CD PDFs")
    ap.add_argument("--addenda", required=False, default=None, help="Folder containing addenda PDFs (optional)")
    ap.add_argument("--out", required=True, help="Output xlsx path")
    ap.add_argument("--config", required=True, help="config.yaml path")
    args = ap.parse_args()

    config = load_config(args.config)

    gmp = ingest_set(config, "GMP", args.gmp)
    bid = ingest_set(config, "BID", args.bid)

    inv_rows = build_inventory_changes(gmp, bid)
    change_rows = compare_sets(config, gmp, bid)

    # Addenda support: compare GMP->Addendum, then Apply Addendum->BID (optional quick view)
    if args.addenda:
        add_pdfs = list_pdfs(args.addenda)
        for pdf in add_pdfs:
            # treat each addendum pdf as its own "set" (keeps attribution clean)
            add_name = Path(pdf).stem
            add_set = DocSet(name=add_name, root=str(Path(pdf).parent), pages=extract_pdf_pages(config, pdf))
            # Compare GMP -> Addendum
            change_rows.extend(compare_sets(config, gmp, add_set))
            # Compare Addendum -> BID (sometimes helps see what stuck)
            change_rows.extend(compare_sets(config, add_set, bid))

    # De-dupe by change_id
    uniq: Dict[str, ChangeRow] = {}
    for r in change_rows:
        uniq[r.change_id] = r
    change_rows = list(uniq.values())

    out_path = args.out
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    write_excel(config, out_path, change_rows, inv_rows)

    print(f"Wrote: {out_path}")
    print(f"Change rows: {len(change_rows)} | Inventory rows: {len(inv_rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
