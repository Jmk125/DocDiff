from __future__ import annotations

from typing import Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import ChangeRow, MatchResult

CHANGE_QUEUE_HEADERS = [
    "Change_ID", "Set_From", "Set_To", "Discipline", "Doc_Type", "Reference",
    "Change_Type", "Change_Summary", "Before_Snippet", "After_Snippet", "Confidence",
    "Auto_Flags", "Impact_Score", "Impact_Rationale", "Estimator_Significance_1to5",
    "Disposition", "Notes",
]


def _autosize(ws: Worksheet, max_width: int = 60) -> None:
    for i, col in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=0)
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, max_width)


def _format(ws: Worksheet, headers: List[str]) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    wrap_cols = {"Before_Snippet", "After_Snippet", "Change_Summary", "Impact_Rationale"}
    for idx, header in enumerate(headers, start=1):
        if header in wrap_cols:
            for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
                row[0].alignment = Alignment(wrap_text=True, vertical="top")
    _autosize(ws)


def write_workbook(path: str, changes: Iterable[ChangeRow], inventory: Iterable[ChangeRow], matches: Iterable[MatchResult]) -> None:
    wb = Workbook()
    cq = wb.active
    cq.title = "Change_Queue"
    cq.append(CHANGE_QUEUE_HEADERS)
    for row in changes:
        cq.append([
            row.change_id, row.set_from, row.set_to, row.discipline, row.doc_type, row.reference,
            row.change_type, row.change_summary, row.before_snippet, row.after_snippet, row.confidence,
            row.auto_flags, row.impact_score, row.impact_rationale, "", "", "",
        ])

    inv = wb.create_sheet("Sheets_Inventory")
    inv.append(CHANGE_QUEUE_HEADERS)
    for row in inventory:
        inv.append([
            row.change_id, row.set_from, row.set_to, row.discipline, row.doc_type, row.reference,
            row.change_type, row.change_summary, row.before_snippet, row.after_snippet, row.confidence,
            row.auto_flags, row.impact_score, row.impact_rationale, "", "", "",
        ])

    matching = wb.create_sheet("Matching")
    matching_headers = ["Set_From", "Reference_From", "Set_To", "Reference_To", "Score", "Confidence", "Reasons"]
    matching.append(matching_headers)
    for result in matches:
        matching.append([
            "", result.from_page.sheet_id or f"p{result.from_page.page_num+1}", "",
            (result.to_page.sheet_id if result.to_page else ""),
            round(result.score, 2), result.confidence, "; ".join(result.reasons),
        ])

    wb.create_sheet("Spec_Inventory")
    wb.create_sheet("Table_Diffs")

    _format(cq, CHANGE_QUEUE_HEADERS)
    _format(inv, CHANGE_QUEUE_HEADERS)
    _format(matching, matching_headers)
    wb.save(path)
